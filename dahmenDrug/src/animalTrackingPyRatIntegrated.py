"""
Created on Tue Mar 21 15:20:43 2023
@author: J02459
"""

import requests
from requests.auth import HTTPBasicAuth

import argparse
from tabula.io import read_pdf
import os
import re
import pandas as pd 
import numpy as np
import time
import openpyxl as op
from openpyxl.styles import Font,PatternFill,Border,Side

baseURL = 'https://pyrat.med.uni-jena.de/pyrat/api/v2/'
url2 = "https://pyrat.med.uni-jena.de/pyrat/api/v2/animals?_key=sacrifice_comment&_key=sacrifice_reason_name&_key=datesacrificed&_key=dateborn&_key=species_name&_key=labid&_key=mutations&_key=sex&_key=animalid&_key=cagenumber&_key=eartag_or_id&_key=licence_number&_key=classification&_key=strain_name&_key=licence_title&_key=owner_fullname&_key=responsible_fullname&_limit=1&eartag=EXNON-81470&state=live&state=sacrificed&state=exported"
response = requests.get(url2, auth = HTTPBasicAuth('1-WUYxhyPozsvRDPCOksFcC9', '11-URrzRso2vNKPvYg5WxZtwT'))
print(response.content)
data = response.json()
data = {k: [v] for k, v in data[0].items()}
df = pd.DataFrame.from_dict(data)
print(df)

Df = pd.read_excel('OP-record-NEW template.xlsx',header=None)
Df.fillna('',inplace=True) 


def dataProcFormat(content):
#Time, bpm
    E = []
    count = 1
    for data in content:
    
        tim = ['']
        bpm = ['']
        ref = ['']
        vap = ['']
        temp = [''] 
        warm = ['40C']
        for i in range(len(list(data[5]))):
    
            if re.findall(r"bpm\b",data[5][i]):
                bpm.append(data[5][i][0:len(data[5][i])-4])
                ref.append('negative')
                warm.append('40C')
            
            if re.findall(r"C\b",data[5][i]):
                temp.append(data[5][i])
                tim.append(data[2][i])    
            if re.findall(r"Dosierun\B",data[3][i]):
                vap.append(str(data[3][i])[34:len(str(data[3][i]))])
                if str(data[3][i])[34:len(str(data[3][i]))]=='4%':
                    tim.append(data[2][i])
        print(len(bpm)-len(vap))
        for t in range(len(bpm)-len(vap)-1):
            if t == len(bpm)-len(vap)-1:
                vap.append('0%')
            else:
                vap.append(vap[-1])
        vap.append('0%')
        if len(bpm)!=10 and len(ref)!=10 and len(warm)!=10:
            for j in range(10-len(bpm)):
                bpm.append('')
                ref.append('')
                warm.append('')
        if len(vap)!=10:
            for k in range(10-len(vap)):
                vap.append('')
        if len(tim)!=10:
            for k in range(10-len(tim)):
                tim.append('')
        
        if len(temp)!=10:
            for k in range(10-len(temp)):
                temp.append('')
        E.append(np.array([tim,bpm, ref, vap, temp, warm],dtype=object))
    
        Df = pd.read_excel('OP-record-NEW template.xlsx',header=None)
        Df.fillna('',inplace=True) 
        
        Df.iloc[8,2]=str(data.iloc[0,2])+str(data.iloc[0,3])+str(data.iloc[0,4])
        Df.iloc[2,5]=str(data.iloc[1,4])[2: len(str(data.iloc[1,4]))]
        Df.iloc[4,2]=str(data.iloc[3,0])[4:len(str(data.iloc[3,0]))]+str(data.iloc[3,1])
        Df.iloc[4,5]=str(data.iloc[3,2])[26:len(str(data.iloc[3,2]))]
        Df.iloc[2,2]= str(data.iloc[1,0])[8:len(str(data.iloc[1,0]))]+str(data.iloc[1,1])+str(data.iloc[1,2])[0:len(str(data.iloc[1,2]))-1]
        Df.iloc[18,1] = 'BPM' 
        Df.iloc[17:23,2:12] = np.array([tim,bpm, ref, vap, temp, warm],dtype=object)
        Df.to_excel(f'{count}.xlsx',encoding = 'latin',index=False,header=None)
        count+=1
    
    wb = op.load_workbook('1.xlsx')
    ws = wb['Sheet1']
    
    my_border = Side(border_style="thin", color="000000")
    my_border1 = Side(border_style="thick", color="000000")
    for cell in range(1,len(list(ws))+1):
        for k in range(1,ws.max_column):
            ws[cell][k].font = Font(bold=True, size=14)
            
            if cell ==3 or cell ==5:
                ws[cell][k].font = Font(bold=True, size=20)
                ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                ws[cell][k].fill = PatternFill('solid', start_color="F2F2F2")
                
            if cell==7 or cell==9:
                if k in [1,2,3,4,5,6]:
                    ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                    ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB")
                
            if cell==10 or cell==11 or cell==12:
                if k in [1,2,3,4,5,6]:
                    ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                    ws[cell][k].fill = PatternFill('solid', start_color="D8E4BC")
                    
            if cell==15 or cell==16:
                ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                if k in [1,2,3,4,5]: 
                    ws[cell][k].fill = PatternFill('solid', start_color="FFFFCC")
            if cell in [m for m in range(18,27)]:
    
                ws[cell][k].border = Border(top=my_border1, left=my_border1, right=my_border1, bottom=my_border1)
                if k==1:
                    if cell%2==0:
                        ws[cell][k].fill = PatternFill('solid', start_color="E6B8B7")
                    else:
                        ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB")
                
                if cell==18:
                    if k in [m for m in range(2,15)]:
                       ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB") 
                    if k==15:
                        ws[cell][k].fill = PatternFill('solid', start_color="D9D9D9")
                else:
                    if cell==19 or cell==20:
                        if k==15:
                            ws[cell][k].fill = PatternFill('solid', start_color="FFFF00")
                    else:
                        if k==15:
                            if cell%2==1:
                                ws[cell][k].fill = PatternFill('solid', start_color="FFFF00")
                            else:
                                ws[cell][k].fill = PatternFill('solid', start_color="F44A4A")
            
            if cell==29:
                ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                ws[cell][k].fill = PatternFill('solid', start_color="FFFF00")
            
            if cell==31:
                ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                ws[cell][k].fill = PatternFill('solid', start_color="F2F2F2")
            
            if cell==52 or cell==53:
                ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB")
    wb.save('1.xlsx')
    
'''from openpyxl.utils import get_column_letter
m = None
main_widths = []
for kk in range(1,ws.max_column):
    column_widths = []
    for celll in range(1,len(list(ws))+1):
        print(kk,celll,str(ws[cell][kk]))
        if str(ws[celll][kk].value)!='None':
            if len(str(ws[celll][kk].value)) > 0 or m==None:
                    m = len(str(ws[celll][kk].value))
                    column_widths += [m]
            else:
                column_widths += [len(str(ws[celll][kk].value))]
        else:
            column_widths += [0]
    main_widths.append(max(column_widths))
    
for i, column_width in enumerate(main_widths,1): 
    ws.column_dimensions[get_column_letter(i)].width = column_width'''
# dfs = read_pdf(f'C:\\Users\\{os.getlogin()}\\.spyder-py3\\'+'Reporttrans00182.pdf', guess=False, pages = 'all',stream=True , encoding="latin", columns = [30,87,125,130,385,420,550])
# dff = pd.DataFrame()

# def main(filePath):
#     stime = time.time()
#     content = dataReadPreProc(filePath)
#     dataProcFormat(content)
#     print(time.time()-stime)
    
# if __name__=="__main__":
#    parser = argparse.ArgumentParser()
#    parser.add_argument("-f", '--file', required = True, help = "the name of the file you want to convert")
#    args = parser.parse_args()
#    main(args.file)
   
