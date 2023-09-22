"""
Created on Tue Mar 21 15:20:43 2023
@author: M03593
"""
import argparse
from tabula.io import read_pdf
import os
import re
import pandas as pd 
import numpy as np
import time
import openpyxl as op
from openpyxl.styles import Font,PatternFill,Border,Side
from copy import copy

def dataReadPreProc(filePath):
    dfs = read_pdf(filePath, guess=False, pages = 'all',stream=True , encoding="latin", columns = [30,87,125,130,385,420,550])
    dff = pd.DataFrame()

#To append dataframes from multiple pages
    for j in range(0,len(dfs)):
        cdf = pd.concat([dff,dfs[j]],ignore_index=True,join='outer')
        dff = cdf

#Dropping irrelevant rows
    cdf.columns = [m+1 for m in range(len(cdf.columns))]
    cdf.fillna('',inplace=True)
    count = 0
    for i in cdf[5]:
        count+=1
        if i=='|Projektlei':
            break
        
    for i in range(len(list(cdf[2]))):
        if not re.findall("\ATie", cdf[2][i]) and not re.findall("\AProj", cdf[2][i]) and not re.findall("\AID", cdf[2][i]) \
            and not re.findall("\A20", cdf[2][i]) and not re.findall("\ADat", cdf[2][i]) and not re.findall("\AUnt", cdf[2][i]) \
            and not i==count-1 and not cdf[2][i]=='':
            cdf.drop(i,inplace=True)

#Resetting the rows and column names
#Data Cleaning
    cdf.reset_index(drop=True,inplace=True)
    cdf.drop([1,4,8],axis=1,inplace=True)
    cdf.columns = [m+1 for m in range(len(cdf.columns))]
    cdf.replace('\?','-',regex=True,inplace=True)
    
    #Segmenting the dataframe as per experiments
    K = []
    t = 0
    exp = cdf.iloc[:5,:]
    for i in range(len(list(cdf[3]))):
        edf = pd.DataFrame()
        if re.findall(r"\bEnde des Exp",cdf[3][i]):
            edf = cdf.iloc[t:i+1,:]
            K.append(edf)
            t = i+2
    edf = cdf.iloc[t:i+1,:]
    K.append(edf)

    M = []
    for data in K:
        edf = pd.concat([exp,data],ignore_index=True,join='outer')
        M.append(edf)
    return M

def dataProcFormat(content):
#Time, bpm
    E = []
    count = 1
    for data in content:
        tim = ['x']
        bpm = ['x']
        ref = ['x']
        vap = ['x']
        temp = ['x'] 
        warm = ['40C']
        LLL = ''
        LML = ''
        RML = ''
        RSL = ''
        RIL = ''
        CSL = ''
        CIL = ''
        for i in range(len(list(data[5]))):
            if re.findall(r"bpm\b",data[5][i]):
                bpm.append(data[5][i][0:len(data[5][i])-4])
                ref.append('negative')
                warm.append('40C')
            if re.findall(r"C\b",data[5][i]):
                temp.append(data[5][i])
                tim.append(data[2][i])    
            if re.findall(r"Dosierun\B",data[3][i]):
                vap.append(str(data[3][i])[34:len(str(data[3][i]))])
                if str(data[3][i])[34:len(str(data[3][i]))]=='4%':
                    tim.append(data[2][i])
            if re.findall(r"\AID",data[1][i]):
                iden = i
            if re.findall(r"Leberresektion\b",data[3][i]):
                weight = i
            if re.findall(r"\bLLL",data[5][i]):
                LLL = str(data[5][i])[4:len(str(data[5][i]))]
            if re.findall(r"\bLML",data[5][i]):
                LML = str(data[5][i])[4:len(str(data[5][i]))]
            if re.findall(r"\bRML",data[5][i]):
                RML = str(data[5][i])[4:len(str(data[5][i]))]
            if re.findall(r"\bRSL",data[5][i]):
                RSL = str(data[5][i])[4:len(str(data[5][i]))]
            if re.findall(r"\bRIL",data[5][i]):
                RIL = str(data[5][i])[4:len(str(data[5][i]))]
            if re.findall(r"\bCSL",data[5][i]):
                CSL = str(data[5][i])[4:len(str(data[5][i]))]
            if re.findall(r"\bCIL",data[5][i]):
                CIL = str(data[5][i])[4:len(str(data[5][i]))]
        
        for t in range(len(bpm)-len(vap)-1):
            if t == len(bpm)-len(vap)-1:
                vap.append('0%')
            else:
                vap.append(vap[-1])
        vap.append('0%')
        if len(bpm)!=10 and len(ref)!=10 and len(warm)!=10:
            for j in range(10-len(bpm)):
                bpm.append('x')
                ref.append('x')
                warm.append('x')
        if len(vap)!=10:
            for k in range(10-len(vap)):
                vap.append('x')
        if len(tim)!=10:
            for k in range(10-len(tim)):
                tim.append('x')
        
        if len(temp)!=10:
            for k in range(10-len(temp)):
                temp.append('x')
        E.append(np.array([tim,bpm, ref, vap, temp, warm],dtype=object))
    
        Df = pd.read_excel(r'C:\Users\M03593\OP-record-NEW template.xlsx',header=None)
        Df.fillna('',inplace=True) 
        
        Df.iloc[8,2]=str(data.iloc[0,2])+str(data.iloc[0,3])+str(data.iloc[0,4])
        Df.iloc[9,2] = Df.iloc[8,2]
        Df.iloc[10,2] = 'Smooth Procedure'
        Df.iloc[2,5]=str(data.iloc[1,4])[2: len(str(data.iloc[1,4]))]
        Df.iloc[4,2]=str(data.iloc[iden,0])[4:len(str(data.iloc[iden,0]))]+str(data.iloc[iden,1])
        Df.iloc[4,5]=str(data.iloc[iden,2])[26:len(str(data.iloc[iden,2]))]
        Df.iloc[2,2]= str(data.iloc[1,0])[8:len(str(data.iloc[1,0]))]+str(data.iloc[1,1])+str(data.iloc[1,2])[0:len(str(data.iloc[1,2]))-1]
        Df.iloc[18,1] = 'Resp. Rate BPM'
        Df.iloc[17:23,2:12] = np.array([tim,bpm, ref, vap, temp, warm],dtype=object)
        Df.iloc[2,8]=str(data[1][weight])
        Df.iloc[15,6]=LLL
        Df.iloc[15,7]=LML
        Df.iloc[15,8]=RML
        Df.iloc[15,9]=RSL
        Df.iloc[15,10]=RIL
        Df.iloc[15,11]=CSL
        Df.iloc[15,12]=CIL
        for i in range(weight):
            if re.findall(r"\bScore",data[5][i]):
                Df.iloc[15,3] = str(data[5][i])
                if str(data[5][i])[-1]=='0':
                    Df.iloc[15,4] = 'Umbelastet'
                elif str(data[5][i])[-1]=='1':
                    Df.iloc[15,4] = 'Mild'
                elif str(data[5][i])[-1]=='2':
                    Df.iloc[15,4] = 'Mild'
                elif str(data[5][i])[-1]=='3':
                    Df.iloc[15,4] = 'Moderate'
                else:
                    Df.iloc[15,4] = 'Severe'
            if re.findall(r"Gewicht:\B",data[5][i]):
                Df.iloc[15,2] = str(data[5][i])[9:len(str(data[5][i]))]
                break   
        Df.to_excel(str(Df.iloc[4,2])+'_'+f'{str(count)}.xlsx',encoding = 'latin',index=False,header=None)
        count+=1
        
def formatting(path):
    for file_name in os.listdir(path):
        if file_name.endswith(".xlsx"):
            wb = op.load_workbook(rf'C:\Users\M03593\{file_name}')
            ws = wb['Sheet1']
            my_border = Side(border_style="thin", color="000000")
            my_border1 = Side(border_style="thick", color="000000")
            for cell in range(1,len(list(ws))+1):
                for k in range(1,ws.max_column):
                    ws[cell][k].font = Font(bold=True, size=14)
                    if cell ==3 or cell ==5:
                        ws[cell][k].font = Font(bold=True, size=20)
                        ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                        ws[cell][k].fill = PatternFill('solid', start_color="F2F2F2")

                    if cell==7 or cell==9:
                        if k in [1,2,3,4,5,6]:
                            ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                            ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB")

                    if cell==10 or cell==11 or cell==12:
                        if k in [1,2,3,4,5,6]:
                            ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                            ws[cell][k].fill = PatternFill('solid', start_color="D8E4BC")

                    if cell==15 or cell==16:
                        ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                        if k in [1,2,3,4,5]: 
                            ws[cell][k].fill = PatternFill('solid', start_color="FFFFCC")
                    if cell in [m for m in range(18,27)]:

                        ws[cell][k].border = Border(top=my_border1, left=my_border1, right=my_border1, bottom=my_border1)
                        if k==1:
                            if cell%2==0:
                                ws[cell][k].fill = PatternFill('solid', start_color="E6B8B7")
                            else:
                                ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB")

                        if cell==18:
                            if k in [m for m in range(2,15)]:
                                ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB") 
                            if k==15:
                                ws[cell][k].fill = PatternFill('solid', start_color="D9D9D9")
                        else:
                            if cell==19 or cell==20:
                                if k==15:
                                    ws[cell][k].fill = PatternFill('solid', start_color="FFFF00")
                            else:
                                if k==15:
                                    if cell%2==1:
                                        ws[cell][k].fill = PatternFill('solid', start_color="FFFF00")
                                    else:
                                        ws[cell][k].fill = PatternFill('solid', start_color="F44A4A")
                    if cell==29:
                        ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                        ws[cell][k].fill = PatternFill('solid', start_color="FFFF00")

                    if cell==31:
                        ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                        ws[cell][k].fill = PatternFill('solid', start_color="F2F2F2")

                    if cell==52 or cell==53:
                        ws[cell][k].border = Border(top=my_border, left=my_border, right=my_border, bottom=my_border)
                        ws[cell][k].fill = PatternFill('solid', start_color="F2DCDB")          
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter (e.g., 'A', 'B', 'C')

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)

                        # Check if the cell is part of a merged range
                        if cell.coordinate in ws.merged_cells:
                            merged_range = ws.merged_cells[cell.coordinate]
                            for merged_cell in merged_range.cells:
                                if len(str(merged_cell.value)) > max_length:
                                    max_length = len(merged_cell.value)
                    except:
                        pass
                adjusted_width = max_length*1.2 + 2  # Add some padding and adjust width

                # Set the calculated column width
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(path)
'''if __name__=='__main__':
    path = r'C:/Users/M03593/Downloads/ilovepdf_split-range/'
    for file_name in os.listdir(path):
        if file_name.endswith(".pdf"):
            dataProcFormat(dataReadPreProc(path+file_name))
            formatting(r'C:\\Users\\M03593\\')'''


def main(filePath):
    stime = time.time()
    for file_name in os.listdir(filePath):
        if file_name.endswith(".pdf"):
            dataProcFormat(dataReadPreProc(filePath+file_name))
            formatting(filePath)
    
    print(time.time()-stime)
    
if __name__=="__main__":
   parser = argparse.ArgumentParser()
   parser.add_argument("-f", '--file', required = True, help = "the name of the file you want to convert")
   args = parser.parse_args()
   main(args.file)
